import json
from datetime import date
from openai import OpenAI
from openpyxl import load_workbook
import csv
import params

OPENAI_KEY = params.OPENAI_KEY
client = OpenAI(api_key=OPENAI_KEY)

def extract_doc_structure_xlsx(file_path):
    """
    Extracts a structured representation of an XLSX file.
    Each worksheet is treated as a table. Returns a list of elements with:
      - 'sheet_index', 'row_index', 'col_index', and 'text'.
    """
    elements = []
    wb = load_workbook(file_path, data_only=True)
    for si, sheet in enumerate(wb.worksheets):
        for ri, row in enumerate(sheet.iter_rows(values_only=True)):
            for ci, cell in enumerate(row):
                text = str(cell) if cell is not None else ""
                elements.append({
                    "type": "table",
                    "sheet_index": si,
                    "row_index": ri,
                    "col_index": ci,
                    "text": text
                })
    return elements

def replace_text_in_xlsx(file_path, replacements, output_path):
    """
    Replaces text in an XLSX file based on precise location.
    Each replacement should include 'sheet_index', 'row_index', and 'col_index'.
    Saves the updated workbook to output_path.
    """
    wb = load_workbook(file_path)
    for rep in replacements:
        loc = rep.get("location", {})
        original = rep.get("original", "")
        new_text = rep.get("replacement", "")
        si = loc.get("sheet_index")
        ri = loc.get("row_index")
        ci = loc.get("col_index")
        try:
            sheet = wb.worksheets[si]
            cell = sheet.cell(row=ri+1, column=ci+1)  # openpyxl uses 1-indexing
            if original in str(cell.value):
                cell.value = str(cell.value).replace(original, new_text, 1)
        except IndexError:
            print(f"Invalid location: {loc}")
    wb.save(output_path)
    return output_path

def extract_doc_structure_csv(file_path):
    """
    Extracts a structured representation of a CSV file.
    Treats the CSV as a single table (sheet_index 0).
    Returns a list of elements with 'row_index', 'col_index', and 'text'.
    """
    elements = []
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for ri, row in enumerate(reader):
            for ci, cell in enumerate(row):
                elements.append({
                    "type": "table",
                    "sheet_index": 0,
                    "row_index": ri,
                    "col_index": ci,
                    "text": cell
                })
    return elements

def replace_text_in_csv(file_path, replacements, output_path):
    """
    Replaces text in a CSV file based on precise location.
    Writes the updated CSV to output_path.
    """
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        rows = list(csv.reader(csvfile))
    for rep in replacements:
        loc = rep.get("location", {})
        original = rep.get("original", "")
        new_text = rep.get("replacement", "")
        ri = loc.get("row_index")
        ci = loc.get("col_index")
        try:
            if original in rows[ri][ci]:
                rows[ri][ci] = rows[ri][ci].replace(original, new_text, 1)
        except IndexError:
            print(f"Invalid location: {loc}")
    with open(output_path, "w", newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(rows)
    return output_path

def get_replacements_from_gpt_csv_xlsx(doc_structure, email, company_context, user_context, user_email):
    """
    Uses GPT-4 to analyze the Excel/CSV document structure and determine which fields need updating.
    Returns a JSON list where each object includes:
      - 'location': {{'type': 'table', 'sheet_index': <sheet_index>, 'row_index': <row_index>, 'col_index': <col_index>}}
      - 'original': the exact text snippet to be replaced.
      - 'replacement': the new text.
    """
    doc_structure_json = json.dumps(doc_structure, ensure_ascii=False, indent=2)
    prompt = f"""
        You are given a quote document represented as a list of document elements from an Excel/CSV file. Each element has index information 
        (for a table: 'sheet_index', 'row_index', 'col_index') along with its text content.
        Additionally, consider the following context:

        An email from a potential client:
        {email}

        Company context:
        {company_context}

        User provided context:
        {user_context}

        Today's date:
        {date.today()}

        User's email:
        {user_email}

        The document may be a template or a previously filled-out template. Your task is to analyze the document elements 
        and complete a quote given the provided context. For each field, provide an object with:
          - 'location': include {{'type': 'table', 'sheet_index': <sheet_index>, 'row_index': <row_index>, 'col_index': <col_index>}}.
          - 'original': the exact text snippet to be updated.
          - 'replacement': the new text to use.
        Return your answer as a JSON list.
        Document structure:
        {doc_structure_json}
    """
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are a professional sales assistant helping generate and update quotes."},
            {"role": "user", "content": prompt}
        ]
    )
    reply = response.choices[0].message.content
    
    reply = reply.split("```json")[-1].split("```")[0]
    try:
        replacements = json.loads(reply)
    except Exception as e:
        print("Error parsing GPT-4 response:", e)
        replacements = []
    return replacements